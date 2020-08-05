# -*- coding: utf-8 -*-

import datetime
import Databases
import calendar
import locale
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

locale.setlocale(locale.LC_ALL, "pl_PL")

overdue_measurement_instruments = Databases.MP2().get_inventory_number_overdue_measurement_instruments()
current_measurement_instruments = Databases.MP2().get_inventory_number_current_measurement_instruments()
next_measurement_instruments = Databases.MP2().get_inventory_number_next_measurement_instruments()


thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

pe = Databases.PE()
mp2 = Databases.MP2()
workbook = Workbook()
sheet = workbook.active
sheet.freeze_panes = 'B2'
sheet.title = calendar.month_name[datetime.date.today().month] + " " + str(datetime.date.today().year)

sheet["A1"] = "Kod obiektu"
sheet["B1"] = "Typ przyrządu pomiarowego"
sheet["C1"] = "Numer inwenatrzowy"
sheet["D1"] = "Data kalibracji"
sheet["E1"] = "Osoba odpowiedzialna/zespół"
sheet["F1"] = "Numer seryjny"
sheet["G1"] = "Model"
sheet["H1"] = "Zakres pomiarowy"
sheet["I1"] = "Czasookres kalibracji"


def create_table(inventory_numbers):
    max_row = sheet.max_row
    for i in range(0, len(inventory_numbers)):
        # Kod obiektu
        sheet["A" + str(i + max_row + 1)] = mp2.get_object_code(inventory_numbers[i])
        # Typ przyrządu pomiarowego
        sheet["B" + str(i + max_row + 1)] = pe.get_type_name(inventory_numbers[i])
        # Numer inwentarzowy
        sheet["C" + str(i + max_row + 1)] = inventory_numbers[i]
        # Data kalibracji
        sheet["D" + str(i + max_row + 1)] = mp2.get_calibration_date(inventory_numbers[i]).strftime("%m.%Y")
        # Osoba odpowiedzialna/zespół za dostarczenie przyrządu pomiarowego do kalibracji
        if "wp" in pe.get_status(inventory_numbers[i]):
            responsible_employee = pe.get_responsible_employee(inventory_numbers[i])
        else:
            responsible_employee = mp2.get_team(inventory_numbers[i])
        sheet["E" + str(i + max_row + 1)] = responsible_employee
        # Numer seryjny
        sheet["F" + str(i + max_row + 1)] = pe.get_serial_number(inventory_numbers[i])
        # Model
        sheet["G" + str(i + max_row + 1)] = pe.get_model(inventory_numbers[i])
        # Zakres pomiarowy
        sheet["H" + str(i + max_row + 1)] = pe.get_range(inventory_numbers[i])
        # Czasookres kalibracji
        calibration_period = pe.get_calibration_period(inventory_numbers[i])
        if calibration_period == 1:
            calibration_period = str(calibration_period) + " rok"
        elif calibration_period in range(2, 5):
            calibration_period = str(calibration_period) + " lata"
        elif calibration_period == 5:
            calibration_period = str(calibration_period) + " lat"
        else:
            pass
        sheet["I" + str(i + max_row + 1)] = calibration_period


def fill_color(overdue_inventory_numbers, current_inventory_numbers):
    red = 'FF0000'
    yellow = 'FFFF00'
    green = '00FF00'
    max_row = sheet.max_row
    for column in range(sheet.min_column, sheet.max_column + 1):
        sheet.cell(1, column).border = thin_border
        for row in range(2, max_row + 1):
            sheet.cell(row, column).border = thin_border
            if row <= len(overdue_inventory_numbers) + 1:
                sheet.cell(row, column).fill = PatternFill(start_color=red, end_color=red, fill_type='solid')
            elif row <= len(overdue_inventory_numbers) + len(current_inventory_numbers) + 1:
                sheet.cell(row, column).fill = PatternFill(start_color=yellow, end_color=yellow, fill_type='solid')
            else:
                sheet.cell(row, column).fill = PatternFill(start_color=green, end_color=green, fill_type='solid')


def set_columns_width():
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for column in range(sheet.min_column, sheet.max_column + 1):
        width = 0
        for row in range(sheet.min_row, sheet.max_row + 1):
            value = sheet.cell(row, column).value
            if value is not None:
                if len(value) > width:
                    width = len(value)
        sheet.column_dimensions[columns[column - 1]].width = width + 2


create_table(overdue_measurement_instruments)
create_table(current_measurement_instruments)
create_table(next_measurement_instruments)
fill_color(overdue_measurement_instruments, current_measurement_instruments)

set_columns_width()

report_name = "Raport kalibracyjny PE {} {}.xlsx".format(calendar.month_name[datetime.date.today().month],
                                                         datetime.date.today().year)
workbook.save(filename=report_name)


